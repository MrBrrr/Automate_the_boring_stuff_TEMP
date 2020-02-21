#! python
# tabulates population and number of census tract for each county

import openpyxl, os, pprint

os.chdir(os.path.join(os.path.dirname(os.getcwd()), "automate_online-materials"))

workbook_handler = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = workbook_handler.active
print(sheet.max_row)


class CountyDataClass(object):
    def __init__(self, state, county, population):
        self.tracts = 1
        self.state = state
        self.county = county
        self.population = population

    def add_pop(self, number):
        self.population += number

    def add_census(self):
        self.tracts += 1


CountyDataDictionary = {}  # county : classObj
# CountyDataList = []

for eachRow in tuple(sheet.rows)[1 : sheet.max_row+1]:
    try:
        CountyDataDictionary[str(eachRow[2].value)]
    except KeyError:
        #  1 - State, 2 - County, 3 - Population
        CountyDataDictionary[eachRow[2].value] = CountyDataClass(eachRow[1].value,
                                                                 eachRow[2].value, eachRow[3].value)
        # print(str(eachRow[1].value))
    else:
        CountyDataDictionary[str(eachRow[2].value)].add_census()
        CountyDataDictionary[str(eachRow[2].value)].add_pop(eachRow[3].value)
        """
        print(CountyDataDictionary[str(eachRow[2].value)].state,  ' ',
              CountyDataDictionary[str(eachRow[2].value)].county, ' ',
              str(CountyDataDictionary[str(eachRow[2].value)].tracts), ' ',
              CountyDataDictionary[str(eachRow[2].value)].population)
        """
workbook_handler.close()


for key, value in CountyDataDictionary.items():
    print(value.state.ljust(4, ' '), value.county.ljust(15, ' '),
          'tracts '+str(value.tracts).ljust(10, ' '), 'population '+str(value.population))

print("Writing to a file")
FileHandler = open('census2010.py', 'w')

for key, value in CountyDataDictionary.items():
    # write
    print(value.state.ljust(4, ' '), value.county.ljust(15, ' '),
          'tracts '+str(value.tracts).ljust(10, ' '), 'population '+str(value.population))

