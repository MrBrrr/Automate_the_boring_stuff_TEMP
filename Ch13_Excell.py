#! python

import openpyxl, os
# it's lower: from openpyxl.utils.cell import get_column_letter, column_index_from_string

os.chdir(os.path.join(os.getcwd(), "automate_online-materials"))
print(os.getcwd())

wb = openpyxl.load_workbook("example.xlsx")
print(type(wb))  # <class 'openpyxl.workbook.workbook.Workbook'>

print(wb.sheetnames)

sheet3 = wb["Sheet3"]
print(type(sheet3))  # <class 'openpyxl.worksheet.worksheet.Worksheet'>

sheet1 = wb.active  # active sheet is sheet1
print(sheet1)  # <Worksheet "Sheet1">

print(sheet1['A1'].value)
c = sheet1['B1']  # c
print(type(c))  # <class 'openpyxl.cell.cell.Cell'>
print("Row: " + str(c.row) + ", column: " + str(c.column) + ", value: " + c.value)
print("Cell's coordinates: " + str(c.coordinate) + ", value: " + c.value)

d = sheet1.cell(row=2, column=1)  # d
print(d.value)

print("Max column: " + str(sheet1.max_column) + ", max row: " + str(sheet1.max_row))
print(sheet1.cell(row=sheet1.max_row+1, column=sheet1.max_column+1).value)  # none

# converts from letters to number
print(openpyxl.utils.cell.column_index_from_string('AAA'))

# converts from numbers to letters
print(openpyxl.utils.cell.get_column_letter(1))

from openpyxl.utils.cell import get_column_letter, column_index_from_string
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(400))
print(get_column_letter(sheet1.max_column))

print(column_index_from_string('ABC'))

# range of cells
for RowOfCellsObj in sheet1['A1':'C3']:
    print(RowOfCellsObj)
    for CellObj in RowOfCellsObj:
        print(CellObj.coordinate + ' : ' + str(CellObj.value))

print(tuple(sheet1.columns))
print(tuple(sheet1.columns)[1])  # it takes second one, because it's tuple now

for each_cell in tuple(sheet1.columns)[1]:
    print(each_cell.value)

print(str(tuple(sheet1.columns)[0][3]) + ":)")  # and this takes first/fourth one of tuple
